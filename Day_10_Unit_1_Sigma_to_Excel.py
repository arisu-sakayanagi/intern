import subprocess, argparse, xlsxwriter, yaml, os
# import pandas as pd

from openpyxl import load_workbook as lw
from openpyxl.utils import get_column_letter
# from openpyxl import Workbook

parser = argparse.ArgumentParser()
parser.add_argument('-t', '--target', type=str, help='target')
parser.add_argument('-c', '--config', type=str, help='config')
parser.add_argument('sigma', metavar='sigma', type=str,help='sigma')
parser.add_argument('-out', '--output', type=str, help='output')
args = parser.parse_args()

if args.output:
    sigma = args.sigma

    query = subprocess.check_output('/home/kami/sigma/tools/sigmac -t'+args.target+' -c'+args.config+' '+sigma, shell=True).decode('utf-8')

    file_name = sigma.split('/')[-1]
    with open(sigma, "r") as stream:
        try:
            dict = yaml.safe_load(stream)
            title = dict['title']
            description = dict['description']
            technique = '\n'.join(dict['tags'])
        except yaml.YAMLError as exc:
            print(exc)


    if os.path.isfile(args.output):
        wb = lw(args.output)
        ws = wb.active
        for col in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col)
            max_col_row = len([cell for cell in ws[col_letter] if cell.value])
            # print("Column: {}, Row numbers: {}".format(col_letter, max_col_row))
        ws['A'+str(max_col_row+1)].value = file_name
        ws['B'+str(max_col_row+1)].value = title
        ws['C'+str(max_col_row+1)].value = description
        ws['D'+str(max_col_row+1)].value = technique
        ws['E'+str(max_col_row+1)].value = query
        wb.save(args.output)


    else:
        workbook = xlsxwriter.Workbook(args.output)
        worksheet = workbook.add_worksheet()
        worksheet.write_row(0,0, ['File Name', 'Title', 'Description', 'Technique', 'Query'])
        worksheet.write_row(1,0, [file_name, title, description, technique, query])
        workbook.close()
        
